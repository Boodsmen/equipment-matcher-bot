"""
Integration тесты для полного цикла работы бота.

Тестируют:
- DOCX → текст → OpenAI → требования → matcher → Excel
- Использование всех 759 моделей из БД (не только 200)
- Корректность процента совпадения
- Формат Excel отчета
"""

import asyncio
import json
import os
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

# Для integration тестов нужны все зависимости установлены
pytest_plugins = ("pytest_asyncio",)


class TestFullCycle:
    """
    Integration тесты для полного цикла обработки ТЗ.

    ВАЖНО: Эти тесты требуют:
    - Установленные зависимости (requirements.txt)
    - Настроенную БД PostgreSQL с импортированными моделями
    - OpenAI API ключ (или мок для тестирования без API)
    """

    @pytest.fixture
    def sample_requirements(self):
        """Пример требований из ТЗ (результат OpenAI парсинга)."""
        return {
            "items": [
                {
                    "item_name": "Коммутатор L3",
                    "model_name": None,
                    "category": "Коммутаторы",
                    "required_specs": {
                        "ports_1g_rj45": 24,
                        "ports_10g_sfp_plus": 4,
                        "power_watt": 200,
                        "layer": 3,
                        "poe_support": True,
                    },
                }
            ]
        }

    @pytest.mark.asyncio
    async def test_matcher_uses_all_models_not_limited_to_200(self):
        """
        Проверка что matcher использует все модели из БД, а не только 200.

        Это критический тест для проверки исправления бага с лимитом.
        """
        from services.matcher import find_matching_models
        from database.crud import get_all_models

        # Подготовка: мокируем БД чтобы вернуть > 200 моделей
        mock_models = []
        for i in range(300):  # Создаем 300 моделей
            model = MagicMock()
            model.id = i
            model.model_name = f"Model_{i}"
            model.category = "Коммутаторы"
            model.source_file = f"v{20 + i % 10}.csv"
            model.specifications = {"ports_1g_rj45": 24 + i % 10}
            model.raw_specifications = {}
            mock_models.append(model)

        with patch("services.matcher.get_all_models", new_callable=AsyncMock) as mock_get_all:
            mock_get_all.return_value = mock_models

            requirements = {
                "items": [
                    {
                        "model_name": None,
                        "category": None,  # Поиск по всей БД
                        "required_specs": {"ports_1g_rj45": 24},
                    }
                ]
            }

            result = await find_matching_models(requirements)

            # Проверяем что обработаны ВСЕ 300 моделей
            total_found = result["summary"]["total_models_found"]
            assert total_found == 300, f"Expected 300 models, got {total_found} (limit was not removed!)"

    @pytest.mark.asyncio
    async def test_numeric_comparison_works_after_bugfix(self):
        """
        Проверка что числовые сравнения работают после исправления бага.

        До исправления: req_num не инициализировалась → все числовые сравнения падали
        После исправления: числовые сравнения должны работать корректно
        """
        from services.matcher import compare_spec_values

        # Простые числа
        assert compare_spec_values(24, 24, "ports") is True
        assert compare_spec_values(24, 30, "ports") is True
        assert compare_spec_values(24, 20, "ports") is False

        # Строки с числами
        assert compare_spec_values("24 порта", 24, "ports") is True
        assert compare_spec_values("24 порта", 30, "ports") is True
        assert compare_spec_values("24 порта", 20, "ports") is False

        # Диапазоны
        assert compare_spec_values("10-20", 25, "range") is True
        assert compare_spec_values("10-20", 15, "range") is False

        # Умножение
        assert compare_spec_values("2x4", 10, "calc") is True
        assert compare_spec_values("2x4", 5, "calc") is False

        # allow_lower с допуском 5%
        assert compare_spec_values(200, 195, "power", allow_lower=True) is True
        assert compare_spec_values(200, 180, "power", allow_lower=True) is False

    @pytest.mark.asyncio
    async def test_match_percentage_calculation(self, sample_requirements):
        """
        Проверка корректности вычисления процента совпадения.
        """
        from services.matcher import calculate_match_percentage

        required_specs = sample_requirements["items"][0]["required_specs"]

        # Модель с 100% совпадением
        model_specs_100 = {
            "ports_1g_rj45": 24,
            "ports_10g_sfp_plus": 4,
            "power_watt": 200,
            "layer": 3,
            "poe_support": True,
        }
        result = calculate_match_percentage(required_specs, model_specs_100)
        assert result["match_percentage"] == 100.0
        assert len(result["matched_specs"]) == 5
        assert len(result["missing_specs"]) == 0

        # Модель с 60% совпадением (3 из 5)
        model_specs_60 = {
            "ports_1g_rj45": 24,
            "ports_10g_sfp_plus": 4,
            "power_watt": 200,
            # layer отсутствует
            # poe_support отсутствует
        }
        result = calculate_match_percentage(required_specs, model_specs_60)
        assert result["match_percentage"] == 60.0
        assert len(result["matched_specs"]) == 3
        assert len(result["missing_specs"]) == 2

    def test_excel_report_has_version_column(self, tmp_path):
        """
        Проверка что Excel отчет содержит колонку "Версия".
        """
        from services.excel_generator import generate_report

        # Подготовка тестовых данных
        match_results = {
            "results": [
                {
                    "requirement": {
                        "item_name": "Test Item",
                        "required_specs": {"ports_1g_rj45": 24},
                    },
                    "matches": {
                        "ideal": [
                            {
                                "model_id": 1,
                                "model_name": "MES3710P",
                                "category": "Коммутаторы",
                                "source_file": "MES3710P_v29.csv",
                                "match_percentage": 100.0,
                                "matched_specs": ["ports_1g_rj45"],
                                "missing_specs": [],
                                "different_specs": {},
                                "specifications": {"ports_1g_rj45": 24},
                                "raw_specifications": {},
                            }
                        ],
                        "partial": [],
                        "not_matched": [],
                    },
                }
            ],
            "summary": {
                "total_requirements": 1,
                "total_models_found": 1,
                "ideal_matches": 1,
                "partial_matches": 0,
            },
        }

        requirements = {"items": [{"item_name": "Test Item", "required_specs": {"ports_1g_rj45": 24}}]}

        # Генерация отчета
        output_dir = str(tmp_path)
        file_path = generate_report(requirements, match_results, output_dir=output_dir)

        # Проверка что файл создан
        assert os.path.exists(file_path)

        # Проверка структуры (требует openpyxl)
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path)
            ws = wb["Сводка"]

            # Проверяем заголовки
            headers = [cell.value for cell in ws[1]]
            assert "Версия" in headers, "Column 'Версия' not found in Excel report!"

            # Проверяем что версия отображается правильно
            version_col_idx = headers.index("Версия") + 1
            version_value = ws.cell(row=2, column=version_col_idx).value
            assert version_value == "v29", f"Expected 'v29', got '{version_value}'"

        except ImportError:
            pytest.skip("openpyxl not installed, skipping Excel structure check")

    def test_reverse_mapping_loaded(self):
        """
        Проверка что reverse_normalization_map.json загружается корректно.
        """
        from services.excel_generator import _load_reverse_mapping

        reverse_mapping = _load_reverse_mapping()

        # Проверяем что файл загружен
        assert isinstance(reverse_mapping, dict)
        assert len(reverse_mapping) > 0

        # Проверяем некоторые ключи
        assert "ports_1g_rj45" in reverse_mapping
        assert "power_watt" in reverse_mapping

        # Проверяем что значения читаемые (на русском)
        assert isinstance(reverse_mapping["ports_1g_rj45"], str)
        assert len(reverse_mapping["ports_1g_rj45"]) > 0


class TestDatabaseIntegration:
    """
    Тесты для проверки работы с реальной БД.

    ВАЖНО: Требуют запущенной БД PostgreSQL с импортированными данными.
    """

    @pytest.mark.asyncio
    @pytest.mark.skip(reason="Requires running PostgreSQL database")
    async def test_database_has_759_models(self):
        """
        Проверка что в БД импортированы все 759 моделей.
        """
        from database.crud import get_models_count

        count = await get_models_count()
        assert count >= 759, f"Expected >= 759 models in DB, got {count}"

    @pytest.mark.asyncio
    @pytest.mark.skip(reason="Requires running PostgreSQL database")
    async def test_category_fallback_strategy(self):
        """
        Проверка что fallback стратегия для категорий работает.

        "Коммутаторы" → ["Управляемый", "Неуправляемый", "Промышленный"]
        """
        from services.matcher import find_matching_models

        requirements = {
            "items": [
                {
                    "model_name": None,
                    "category": "Коммутаторы",
                    "required_specs": {"ports_1g_rj45": 24},
                }
            ]
        }

        result = await find_matching_models(requirements)

        # Должны найтись модели из подкатегорий "Управляемый", "Неуправляемый" и т.д.
        total_found = result["summary"]["total_models_found"]
        assert total_found > 0, "No models found with category fallback!"


class TestPerformance:
    """Тесты производительности."""

    @pytest.mark.asyncio
    @pytest.mark.skip(reason="Requires running PostgreSQL database")
    async def test_full_cycle_performance(self):
        """
        Проверка что полный цикл обработки занимает < 10 секунд.
        """
        import time
        from services.matcher import find_matching_models

        requirements = {
            "items": [
                {
                    "model_name": None,
                    "category": None,  # Поиск по всей БД (759 моделей)
                    "required_specs": {"ports_1g_rj45": 24, "power_watt": 200},
                }
            ]
        }

        start_time = time.time()
        result = await find_matching_models(requirements)
        elapsed_time = time.time() - start_time

        assert elapsed_time < 10.0, f"Full cycle took {elapsed_time:.2f}s (expected < 10s)"

        # Проверяем что нашлись модели
        total_found = result["summary"]["total_models_found"]
        assert total_found > 0, "No models found"


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
