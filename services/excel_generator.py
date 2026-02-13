"""
Модуль для генерации Excel отчетов с результатами сопоставления моделей.

Структура отчета (по образцу matching_result_tz_uq91nirt):
- Лист "Сводка"           — метаданные + статистика + топ-10
- Лист "Все совпадения"   — все модели ≥ min_percentage (sorted)
- Лист "Детали совпадений"— построчное сравнение характеристик
- Лист "Не сопоставленные"— позиции без подходящих моделей
"""

import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils.logger import logger

# ─── Цвета ───────────────────────────────────────────────────────────────────
COLOR_GREEN      = "C6EFCE"   # ✅ Совпадает
COLOR_YELLOW     = "FFEB9C"   # ⚠️ Частичное / старые версии
COLOR_RED        = "FFC7CE"   # ❌ Не совпадает
COLOR_ORANGE     = "FFD699"   # ? Нет в каталоге
COLOR_GRAY       = "D9D9D9"   # Заголовок
COLOR_LIGHT_GRAY = "F2F2F2"   # Нечётные строки
COLOR_BLUE_HDR   = "BDD7EE"   # Заголовок секции модели
COLOR_DARK       = "404040"   # Тёмный шрифт

# ─── Reverse mapping (canonical_key → читаемое название) ─────────────────────
_REVERSE_MAPPING_CACHE: Optional[Dict[str, str]] = None


def _load_reverse_mapping() -> Dict[str, str]:
    global _REVERSE_MAPPING_CACHE
    if _REVERSE_MAPPING_CACHE is not None:
        return _REVERSE_MAPPING_CACHE
    try:
        path = Path(__file__).parent.parent / "data" / "reverse_normalization_map.json"
        with open(path, "r", encoding="utf-8") as f:
            _REVERSE_MAPPING_CACHE = json.load(f)
        logger.debug(f"Loaded reverse mapping: {len(_REVERSE_MAPPING_CACHE)} keys")
    except Exception as e:
        logger.warning(f"Failed to load reverse_normalization_map.json: {e}")
        _REVERSE_MAPPING_CACHE = {}
    return _REVERSE_MAPPING_CACHE


def _readable_key(key: str) -> str:
    return _load_reverse_mapping().get(key, key.replace("_", " ").title())


# ─── Версия из source_file ────────────────────────────────────────────────────

def _parse_version(source_file: str) -> str:
    if not source_file:
        return "—"
    m = re.search(r'finalUPDv\.(\d+)\.(\d+)', source_file)
    if m:
        return f"finalUPD v{m.group(1)}.{m.group(2)}"
    if 'finalUPD' in source_file:
        return "finalUPD"
    m = re.search(r'v(\d+)(?:\.(\d+))?', source_file)
    if m:
        v = f"v{m.group(1)}"
        if m.group(2):
            v += f".{m.group(2)}"
        if '_new' in source_file:
            v += " (new)"
        return v
    return source_file


# ─── Вспомогательные функции ─────────────────────────────────────────────────

def _fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _bold(size: int = 11, color: str = None) -> Font:
    kwargs = {"bold": True, "size": size}
    if color:
        kwargs["color"] = color
    return Font(**kwargs)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _auto_width(ws, min_w: int = 8, max_w: int = 60) -> None:
    for col in ws.columns:
        best = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    best = max(best, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(best + 2, min_w), max_w)


def _header_row(ws, row: int, n_cols: int, bg: str = COLOR_GRAY) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _bold()
        cell.fill = _fill(bg)
        cell.alignment = _center()


def _comparison_detail(req_val: Any, mod_val: Any) -> str:
    """Generate a short comparison description like '54.0 >= 32.0'."""
    try:
        from services.matcher import extract_number, extract_number_with_operator
        req_num, op = extract_number_with_operator(req_val)
        mod_num = extract_number(mod_val)
        if req_num is not None and mod_num is not None:
            op_display = op if op != ">=" else ">="
            return f"{mod_num} {op_display} {req_num}"
    except Exception:
        pass
    if isinstance(req_val, str) and isinstance(mod_val, str):
        r, m = req_val.strip().lower(), mod_val.strip().lower()
        if r == m:
            return "Exact text match"
        if r in m:
            return f"'{req_val}' found in '{mod_val}'"
    return ""


# ─── Лист 1: Сводка ──────────────────────────────────────────────────────────

def _create_summary_sheet(
    wb: Workbook,
    match_results: Dict[str, Any],
    filename: str,
    processing_time: float,
    threshold: int,
    min_percentage: float,
) -> None:
    ws = wb.active
    ws.title = "Сводка"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22

    def kv(label: str, value: Any, bold_val: bool = False) -> None:
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=label).font = _bold()
        cell = ws.cell(row=r, column=2, value=value)
        if bold_val:
            cell.font = _bold()

    # ── Заголовок ──
    ws.append(["Результаты сопоставления оборудования"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = _center()
    ws.cell(row=1, column=1).fill = _fill(COLOR_GRAY)
    ws.row_dimensions[1].height = 28

    ws.append([])  # empty row

    # ── Метаданные ──
    kv("Файл ТЗ:", filename)
    kv("Дата обработки:", datetime.now().strftime("%d.%m.%Y %H:%M"))
    if processing_time:
        kv("Время обработки:", f"{processing_time:.2f} сек")

    ws.append([])

    # ── Статистика ──
    results = match_results.get("results", [])
    total_reqs = sum(
        len(r["requirement"].get("required_specs", {}))
        for r in results
    )
    # Collect all matches ≥ min_percentage across all requirements
    all_filtered = []
    for result in results:
        for cat in ("ideal", "partial", "not_matched"):
            for m in result["matches"].get(cat, []):
                if m["match_percentage"] >= min_percentage:
                    all_filtered.append(m)

    kv("Статистика требований:", None)
    r = ws.max_row
    ws.cell(row=r, column=2).value = None

    kv("Всего характеристик:", total_reqs)
    kv("Позиций оборудования:", len(results))
    kv("Найдено моделей (≥80%):", len(all_filtered))
    best_pct = max((m["match_percentage"] for m in all_filtered), default=0.0)
    kv("Лучшее совпадение:", f"{best_pct:.1f}%", bold_val=True)
    kv("Порог отображения:", f"{min_percentage:.0f}%")

    ws.append([])

    # ── Топ-10 ──
    r_hdr = ws.max_row + 1
    ws.append(["Топ-10 совпадений"])
    ws.merge_cells(start_row=r_hdr, start_column=1, end_row=r_hdr, end_column=5)
    ws.cell(row=r_hdr, column=1).font = _bold(12)
    ws.cell(row=r_hdr, column=1).alignment = _left()

    headers = ["№", "Модель", "Совпадение %", "Совпало", "Всего"]
    ws.append(headers)
    _header_row(ws, ws.max_row, len(headers))

    # Collect top models (deduplicated by name — pick best %)
    top_models: Dict[str, Dict] = {}
    for result in results:
        total_specs = len(result["requirement"].get("required_specs", {}))
        for cat in ("ideal", "partial"):
            for m in result["matches"].get(cat, []):
                if m["match_percentage"] >= min_percentage:
                    name = m["model_name"]
                    if name not in top_models or m["match_percentage"] > top_models[name]["match_percentage"]:
                        top_models[name] = {**m, "_total_specs": total_specs}

    sorted_top = sorted(top_models.values(), key=lambda x: x["match_percentage"], reverse=True)[:10]

    for i, m in enumerate(sorted_top, 1):
        pct = m["match_percentage"]
        matched = len(m.get("matched_specs", []))
        total = m["_total_specs"]
        ws.append([i, m["model_name"], f"{pct:.1f}%", matched, total])
        r_cur = ws.max_row
        # Color by percentage
        if pct == 100.0:
            bg = COLOR_GREEN
        elif pct >= threshold:
            bg = COLOR_YELLOW
        else:
            bg = COLOR_ORANGE
        for c in range(1, 6):
            ws.cell(row=r_cur, column=c).fill = _fill(bg)
            ws.cell(row=r_cur, column=c).alignment = _center()

    ws.auto_filter.ref = f"A{r_hdr + 1}:E{ws.max_row}"
    logger.info(f"Summary sheet created: top {len(sorted_top)} models")


# ─── Лист 2: Все совпадения ──────────────────────────────────────────────────

def _create_all_matches_sheet(
    wb: Workbook,
    match_results: Dict[str, Any],
    threshold: int,
    min_percentage: float,
) -> None:
    ws = wb.create_sheet("Все совпадения")

    headers = ["№", "Модель оборудования", "Позиция ТЗ", "Совпадение %",
               "Совпало требований", "Не совпало", "Не сопоставлено"]
    ws.append(headers)
    _header_row(ws, 1, len(headers))

    row_num = 1
    for result in match_results.get("results", []):
        req = result["requirement"]
        req_name = req.get("item_name") or req.get("model_name") or "—"
        total_specs = len(req.get("required_specs", {}))

        for cat in ("ideal", "partial", "not_matched"):
            for m in result["matches"].get(cat, []):
                pct = m["match_percentage"]
                if pct < min_percentage:
                    continue
                matched = len(m.get("matched_specs", []))
                different = len(m.get("different_specs", {}))
                unmapped = len(m.get("unmapped_specs", m.get("missing_specs", [])))

                ws.append([
                    row_num,
                    m["model_name"],
                    req_name,
                    f"{pct:.1f}%",
                    matched,
                    different,
                    unmapped,
                ])
                r = ws.max_row
                if pct == 100.0:
                    bg = COLOR_GREEN
                elif pct >= threshold:
                    bg = COLOR_YELLOW
                else:
                    bg = COLOR_ORANGE
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = _fill(bg)
                    ws.cell(row=r, column=c).alignment = _center()
                row_num += 1

    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)
    logger.info(f"All-matches sheet: {row_num - 1} rows")


# ─── Лист 3: Детали совпадений ───────────────────────────────────────────────

def _create_details_sheet(
    wb: Workbook,
    match_results: Dict[str, Any],
    threshold: int,
    min_percentage: float,
    max_models: int = 50,
) -> None:
    ws = wb.create_sheet("Детали совпадений")
    reverse_mapping = _load_reverse_mapping()

    # Fixed column widths
    ws.column_dimensions["A"].width = 6    # Статус
    ws.column_dimensions["B"].width = 42   # Характеристика
    ws.column_dimensions["C"].width = 22   # Требуется
    ws.column_dimensions["D"].width = 22   # Фактически
    ws.column_dimensions["E"].width = 30   # Детали

    current_row = 1
    model_counter = 0

    for result in match_results.get("results", []):
        req = result["requirement"]
        req_name = req.get("item_name") or req.get("model_name") or "—"
        required_specs = req.get("required_specs", {})

        for cat in ("ideal", "partial"):
            for m in result["matches"].get(cat, []):
                pct = m["match_percentage"]
                if pct < min_percentage:
                    continue
                if model_counter >= max_models:
                    break

                model_counter += 1
                version = _parse_version(m.get("source_file", ""))
                matched_specs = set(m.get("matched_specs", []))
                unmapped_specs = set(m.get("unmapped_specs", m.get("missing_specs", [])))
                different_specs = m.get("different_specs", {})
                model_specs = m.get("specifications") or {}

                # ── Секция-заголовок модели ──
                header_text = f"{model_counter}. {m['model_name']} ({req_name}) — {pct:.1f}%"
                ws.cell(row=current_row, column=1, value=header_text)
                ws.merge_cells(
                    start_row=current_row, start_column=1,
                    end_row=current_row, end_column=5,
                )
                cell = ws.cell(row=current_row, column=1)
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = _fill("2E75B6")
                cell.alignment = _left()
                ws.row_dimensions[current_row].height = 20
                current_row += 1

                # ── Заголовки колонок ──
                for ci, hdr in enumerate(["Статус", "Характеристика", "Требуется", "Фактически", "Детали"], 1):
                    c = ws.cell(row=current_row, column=ci, value=hdr)
                    c.font = _bold()
                    c.fill = _fill(COLOR_GRAY)
                    c.alignment = _center()
                current_row += 1

                # ── Строки характеристик ──
                for spec_i, (key, req_val) in enumerate(required_specs.items()):
                    readable = reverse_mapping.get(key, key.replace("_", " ").title())
                    mod_val = model_specs.get(key)

                    if key in matched_specs:
                        status = "✓"
                        bg = COLOR_GREEN
                        detail = _comparison_detail(req_val, mod_val)
                    elif key in unmapped_specs:
                        status = "?"
                        bg = COLOR_ORANGE
                        mod_val = "—"
                        detail = "Нет в каталоге"
                    elif key in different_specs:
                        status = "✗"
                        bg = COLOR_RED
                        detail = _comparison_detail(req_val, mod_val)
                    else:
                        status = "—"
                        bg = None
                        detail = ""

                    row_bg = bg or (COLOR_LIGHT_GRAY if spec_i % 2 == 0 else None)

                    values = [
                        status,
                        readable,
                        str(req_val) if req_val is not None else "—",
                        str(mod_val) if mod_val is not None else "—",
                        detail,
                    ]
                    for ci, v in enumerate(values, 1):
                        cell = ws.cell(row=current_row, column=ci, value=v)
                        if row_bg:
                            cell.fill = _fill(row_bg)
                        cell.alignment = _left() if ci > 1 else _center()
                    current_row += 1

                # Empty separator row
                current_row += 1

            if model_counter >= max_models:
                break

    logger.info(f"Details sheet: {model_counter} models, {current_row - 1} rows")


# ─── Лист 4: Не сопоставленные ───────────────────────────────────────────────

def _create_unmatched_sheet(
    wb: Workbook,
    match_results: Dict[str, Any],
    min_percentage: float,
) -> None:
    ws = wb.create_sheet("Не сопоставленные")

    has_unmatched = False
    for result in match_results.get("results", []):
        matches = result["matches"]
        all_above = any(
            m["match_percentage"] >= min_percentage
            for cat in ("ideal", "partial", "not_matched")
            for m in matches.get(cat, [])
        )
        if not all_above:
            has_unmatched = True
            break

    if not has_unmatched:
        ws.append(["Все требования сопоставлены с базой данных"])
        ws.cell(row=1, column=1).font = _bold()
        ws.cell(row=1, column=1).fill = _fill(COLOR_GREEN)
        return

    headers = ["№", "Позиция ТЗ", "Категория", "Лучшее совпадение %", "Характеристик"]
    ws.append(headers)
    _header_row(ws, 1, len(headers))

    row_num = 1
    for result in match_results.get("results", []):
        req = result["requirement"]
        matches = result["matches"]
        req_name = req.get("item_name") or req.get("model_name") or "—"
        category = req.get("category") or "—"
        n_specs = len(req.get("required_specs", {}))

        all_models = (
            matches.get("ideal", [])
            + matches.get("partial", [])
            + matches.get("not_matched", [])
        )
        best = max((m["match_percentage"] for m in all_models), default=0.0)

        if best < min_percentage:
            ws.append([row_num, req_name, category, f"{best:.1f}%", n_specs])
            r = ws.max_row
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = _fill(COLOR_RED if best == 0 else COLOR_ORANGE)
                ws.cell(row=r, column=c).alignment = _center()
            row_num += 1

    _auto_width(ws)


# ─── Публичная функция ────────────────────────────────────────────────────────

def generate_report(
    requirements: Dict[str, Any],
    match_results: Dict[str, Any],
    output_dir: str = "temp_files",
    threshold: int = 70,
    min_percentage: float = 80.0,
    filename: str = "",
    processing_time: float = 0.0,
) -> str:
    """
    Генерация Excel отчёта с результатами сопоставления.

    Структура файла:
    - Сводка             — метаданные + статистика + топ-10
    - Все совпадения     — плоский список всех моделей ≥ min_percentage
    - Детали совпадений  — построчное сравнение (топ-50 моделей)
    - Не сопоставленные  — позиции без подходящих моделей
    """
    logger.info(f"Generating Excel report (min={min_percentage}%, threshold={threshold}%)…")

    wb = Workbook()

    _create_summary_sheet(wb, match_results, filename, processing_time, threshold, min_percentage)
    _create_all_matches_sheet(wb, match_results, threshold, min_percentage)
    _create_details_sheet(wb, match_results, threshold, min_percentage)
    _create_unmatched_sheet(wb, match_results, min_percentage)

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(output_dir, f"tender_match_report_{timestamp}.xlsx")
    wb.save(file_path)

    logger.info(f"Excel report saved: {file_path} ({len(wb.sheetnames)} sheets)")
    return file_path
